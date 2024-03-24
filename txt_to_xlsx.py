from TkinterMaster import Tk
from tkinter.filedialog import askopenfilename, askdirectory
import tkinter as tk
from openpyxl import Workbook
from sys import exit

Tk().withdraw()
tk.messagebox.showinfo("После завтрак", "Выберите txt после завтрака")

Tk().withdraw()
breakfast = askopenfilename()

tk.messagebox.showinfo("После Обеда", "Выберите txt после обеда")

Tk().withdraw()
dinner = askopenfilename()

tk.messagebox.showinfo("Сохранение", "Выберите место сохранения итоговых файлов")

Tk().withdraw()
dir = askdirectory()

# Read the text file
try:
    with open(breakfast, 'r') as file:
        lines = file.readlines()
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся открыть файл")
    exit()
# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Write the text file contents to the workbook
for row, line in enumerate(lines, start=1):
    columns = line.strip().split('&')
    for i in range(len(columns)):
        if columns[i] == "1":
            columns[i] = 1
        if columns[i] == "0":
            columns[i] = 0
    for col, value in enumerate(columns, start=1):
        sheet.cell(row=row, column=col, value=value)

try:
    workbook.save(dir + '/breakfast.xlsx')
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся записать файл")
    exit()

try:
    with open(dinner, 'r') as file:
        lines = file.readlines()
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся открыть файл")
    exit()

# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Write the text file contents to the workbook
for row, line in enumerate(lines, start=1):
    columns = line.strip().split('&')
    for i in range(len(columns)):
        if columns[i] == "1":
            columns[i] = 1
        if columns[i] == "0":
            columns[i] = 0
    for col, value in enumerate(columns, start=1):
        sheet.cell(row=row, column=col, value=value)

try:
    workbook.save(dir + '/dinner.xlsx')
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся записать файл")
    exit()