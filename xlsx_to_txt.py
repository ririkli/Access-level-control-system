from TkinterMaster import Tk
from tkinter.filedialog import askopenfilename, askdirectory
import tkinter as tk
import openpyxl
from sys import exit

Tk().withdraw()
tk.messagebox.showinfo("Завтрак", "Выберите таблицу завтракающих")

Tk().withdraw()
breakfast = askopenfilename()

tk.messagebox.showinfo("Обед", "Выберите таблицу обедающих")

Tk().withdraw()
dinner = askopenfilename()

tk.messagebox.showinfo("Сохранение", "Выберите место сохранения итоговых файлов")

Tk().withdraw()
dir = askdirectory()

try:
    wookbook = openpyxl.load_workbook(breakfast)
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся открыть файл")
    exit()

worksheet = wookbook.active
s=""

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        s += str(col[i].value)+"&"
    s += "\n"
try:
    with open(dir + '/breakfast.txt', 'w') as f:
        f.write(s)
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся записать файл")
    exit()

try:
    wookbook = openpyxl.load_workbook(dinner)
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся открыть файл")
    exit()

worksheet = wookbook.active
s=""

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        s += str(col[i].value)+"&"
    s += "\n"
try:
    with open(dir + '/dinner.txt', 'w') as f:
        f.write(s)
except:
    tk.messagebox.showinfo("Ошибка", "Не удаётся записать файл")
    exit()